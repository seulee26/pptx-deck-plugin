"""Generate a synthetic sample.xlsx exercising every item type (for smoke testing)."""
from pathlib import Path
from openpyxl import Workbook

OUT = Path(__file__).resolve().parent.parent / "out" / "sample.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
wb.remove(wb.active)

meta = wb.create_sheet("meta")
meta.append(["title", "2026 Business Review"])
meta.append(["subtitle", "Q1 performance and outlook"])
meta.append(["company", "Acme Corp"])
meta.append(["closing", "감사합니다"])

perf = wb.create_sheet("Performance")
perf.append(["KPI", "Revenue", "Total Q1 revenue", "₩12.4B", "+18%"])
perf.append(["KPI", "Active users", "MAU this quarter", "2,340,000", "+12%"])
perf.append(["KPI", "NPS", "Net promoter score", "72%", "+4"])

growth = wb.create_sheet("Growth")
growth.append(["PROCESS", "Onboarding flow", "", "Sign up", "Verify", "Activate", "Use"])
growth.append(["MATRIX", "SWOT", "", "Strong brand", "Legacy stack", "AI adoption", "New entrants"])
growth.append(["CONTENT", "North Star", "Deliver the most reliable platform in our segment by end of year."])

ops = wb.create_sheet("Operations")
ops.append(["TABLE", "Monthly KPIs", "", "Month", "Revenue", "Users"])
ops.append(["",      "",              "",  "Jan",   "₩3.8B",   "1.9M"])
ops.append(["",      "",              "",  "Feb",   "₩4.1B",   "2.1M"])
ops.append(["",      "",              "",  "Mar",   "₩4.5B",   "2.3M"])
ops.append(["IMAGE", "Product shots", "Hero screenshots across devices"])

wb.save(OUT)
print(f"wrote {OUT}")
